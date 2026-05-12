import os
import json
import time
from io import BytesIO
from pathlib import Path
from flask import Flask, request, jsonify, render_template, send_file
from dotenv import load_dotenv
from google import genai
from google.genai import types
from google.genai.errors import ServerError
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 로컬: .env 또는 ../gemini_setup/.env 에서 로드 / Vercel: 환경변수에서 직접 로드
load_dotenv()
load_dotenv(Path(__file__).parent.parent / "gemini_setup" / ".env")

api_key = os.getenv("GEMINI_API_KEY")
if not api_key:
    raise ValueError("GEMINI_API_KEY를 찾을 수 없습니다.")

client = genai.Client(api_key=api_key)
MODELS = ["gemini-2.5-flash", "gemini-2.0-flash", "gemini-2.0-flash-lite"]

PROMPT = """이 영수증 이미지를 분석하여 아래 JSON 형식으로만 응답해주세요.

{
  "가맹점명": "상점 이름 (인식 불가 시 '알 수 없음')",
  "결제일시": "YYYY-MM-DD HH:MM 형식 (인식 불가 시 '알 수 없음')",
  "총금액": 숫자만 (원 단위 정수, 인식 불가 시 0),
  "항목": [
    {"품목명": "품목 이름", "수량": 숫자, "단가": 숫자, "금액": 숫자}
  ]
}

주의사항:
- 모든 금액은 숫자만 기입 (₩, 원, 쉼표 제외)
- 수량·단가를 알 수 없는 항목은 수량 1, 단가에 소계 금액 기입
- 총금액이 영수증에 명시된 경우 반드시 그 값 사용
- 영수증이 아닌 이미지는 모든 필드를 '알 수 없음' 또는 0으로 반환
- 반드시 유효한 JSON만 출력, 다른 설명 없음"""

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/analyze", methods=["POST"])
def analyze():
    if "image" not in request.files:
        return jsonify({"error": "이미지 파일이 없습니다."}), 400

    file = request.files["image"]
    if not file.filename:
        return jsonify({"error": "파일이 선택되지 않았습니다."}), 400

    mime_type = file.content_type or "image/jpeg"
    if not mime_type.startswith("image/"):
        return jsonify({"error": "이미지 파일만 업로드 가능합니다."}), 400

    image_bytes = file.read()
    last_error = None

    for model in MODELS:
        for attempt in range(3):
            try:
                response = client.models.generate_content(
                    model=model,
                    contents=[
                        types.Part.from_bytes(data=image_bytes, mime_type=mime_type),
                        PROMPT,
                    ],
                    config=types.GenerateContentConfig(
                        response_mime_type="application/json"
                    ),
                )
                # BOM(﻿) 및 앞뒤 공백 제거 후 파싱
                text = response.text.lstrip('﻿').strip()
                data = json.loads(text)
                return jsonify({"success": True, "data": data, "model": model})

            except ServerError as e:
                last_error = e
                if attempt < 2:
                    time.sleep(2 ** attempt)
                continue
            except json.JSONDecodeError:
                return jsonify({"error": "JSON 파싱 실패", "raw": response.text}), 500
            except Exception as e:
                return jsonify({"error": str(e)}), 500

    return jsonify({"error": f"모든 모델이 응답하지 않습니다. 잠시 후 다시 시도해주세요. ({last_error})"}), 503


@app.route("/export", methods=["POST"])
def export_excel():
    results = request.get_json()
    if not results:
        return jsonify({"error": "데이터가 없습니다."}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "영수증 분석"

    header_fill = PatternFill("solid", fgColor="4F6EF7")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="EEF1FF")
    center      = Alignment(horizontal="center", vertical="center")
    right       = Alignment(horizontal="right",  vertical="center")

    headers = ["파일명", "가맹점명", "결제일시", "품목명", "수량", "단가(원)", "금액(원)", "총금액(원)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    row_idx = 2
    for i, result in enumerate(results):
        filename = result.get("filename", "")
        data     = result.get("data", {})
        items    = data.get("항목", [])
        fill     = alt_fill if i % 2 == 0 else PatternFill()

        rows_to_write = items if items else [{}]
        for j, item in enumerate(rows_to_write):
            ws.append([
                filename                       if j == 0 else "",
                data.get("가맹점명", "")       if j == 0 else "",
                data.get("결제일시", "")       if j == 0 else "",
                item.get("품목명", "항목 없음"),
                item.get("수량", ""),
                item.get("단가", ""),
                item.get("금액", ""),
                data.get("총금액", 0)          if j == 0 else "",
            ])
            for col, cell in enumerate(ws[row_idx], 1):
                cell.fill = fill
                cell.alignment = right if col >= 5 else Alignment(vertical="center")
            row_idx += 1

    col_widths = [22, 18, 18, 24, 8, 14, 14, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="영수증_분석결과.xlsx",
    )


if __name__ == "__main__":
    print("서버 시작: http://localhost:5000")
    app.run(debug=True, port=5000)
